print('5--------------------')
import json

data = {
    "name": "홍길동",
    "age": 25,
    "city": "서울"
}

with open("data.json", "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=4)

print("data.json 저장 완료")

print('6--------------------')
import os

folder = "sample_folder"

if not os.path.exists(folder):
    print(f"폴더가 없습니다: {folder}")
else:
    items = os.listdir(folder)
    print(f"[{folder}] 목록:")
    for item in items:
        path = os.path.join(folder, item)
        if os.path.isdir(path):
            print(f"[폴더] {item}")
        else:
            print(f"[파일] {item}")


with open("data.json", "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=4)

print("data.json 저장 완료")

print('7--------------------')
from pathlib import Path

folder = Path("sample_folder")

if not folder.exists():
    print(f"폴더가 없습니다: {folder}")
else:
    txt_files = list(folder.glob("*.txt"))
    print(f"[{folder}] .txt 파일 목록:")
    for f in txt_files:
        print(f.name)

print('8--------------------')
from pathlib import Path

new_folder = Path("new_folder")
new_folder.mkdir(exist_ok=True)

print("new_folder 준비 완료")

print('9--------------------')
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

wb.save("data.xlsx")
print("data.xlsx 생성 완료")

print('10-------------------')
from openpyxl import load_workbook

wb = load_workbook("data.xlsx")
ws = wb.worksheets[0]  # 첫 번째 시트

# 예시로 데이터 추가
ws["A1"] = "이름"
ws["B1"] = "나이"
ws["C1"] = "도시"

ws.append(["홍길동", 25, "서울"])
ws.append(["김철수", 30, "부산"])

wb.save("data.xlsx")
print("data.xlsx에 데이터 추가 완료")
